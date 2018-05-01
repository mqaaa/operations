#!/usr/bin/python
# _*_ coding: UTF-8 _*_
from __future__ import print_function
import openpyxl
import os


def process_worksheet(sheet):
    avg_column = sheet.max_column + 1
    sum_column = sheet.max_column + 2

    for row in sheet.iter_rows(min_row=2, min_col=3):
        scores = [cell.value for cell in row]
        sum_scores = sum(scores)
        avg_scores = sum_column / len(scores)
        sheet.cell(row=row[0].row, column=avg_column).value = avg_scores
        sheet.cell(row=row[0].row, column=sum_column).value = sum_scores
    
    sheet.cell(row=1,column=avg_column).value = 'avg'
    sheet.cell(row=1,column=sum_column).value = 'sum'


def main():
    os.chdir('D://')
    wb = openpyxl.load_workbook('example.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    process_worksheet(sheet)
    wb.save('exampe_copy.xlsx')


if __name__ == '__main__':
    main()

